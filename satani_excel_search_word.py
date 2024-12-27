"""
Excelの行列から特定の文字列を検索し、セル番地を返す
"""

import openpyxl
from openpyxl.utils import coordinate_to_tuple

# 特定の列を検索
def search_column(column, keyword):
    result = []
    for cell in column:
        # セルのデータを文字列に変換
        try:
            value = str(cell.value)
        # 文字列に変換できないデータはスキップ
        except:
            continue
        # キーワードに一致するセルの番地を取得
        if value == keyword:
            cell_address = openpyxl.utils.get_column_letter(cell.column) +  str(cell.row)
            tpl = coordinate_to_tuple(cell_address)
            result.append(tpl)

    return result

# 特定の行を検索
def search_row(row, keyword):
    result = []
    for cell in row:
        # セルのデータを文字列に変換
        try:
            value = str(cell.value)
        # 文字列に変換できないデータはスキップ
        except:
            continue
        # キーワードに一致するセルの番地を取得
        if value == keyword:
            cell_address = openpyxl.utils.get_column_letter(cell.column) +  str(cell.row)
            tpl = coordinate_to_tuple(cell_address)
            result.append(tpl)
            
    return result

# 特定の範囲を検索
def search_rectangle(rectangle, keyword):
    result = []
    for col in rectangle:
        for cell in col:
            # セルのデータを文字列に変換
            try:
                value = str(cell.value)
            # 文字列に変換できないデータはスキップ
            except:
                continue
            # キーワードに一致するセルの番地を取得
            if value == keyword:
                cell_address = openpyxl.utils.get_column_letter(cell.column) +  str(cell.row)
                tpl = coordinate_to_tuple(cell_address)
                result.append(tpl)
            
    return result

# シート全体を検索
def search_entire_sheet(ws, keyword, cordinate="tuple"):
    result = []
    for col in ws.columns:
        for cell in col:
            # セルのデータを文字列に変換
            try:
                value = str(cell.value)
            # 文字列に変換できないデータはスキップ
            except:
                continue
            # キーワードに一致するセルの番地を取得
            if value == keyword:
                cell_address = openpyxl.utils.get_column_letter(cell.column) +  str(cell.row)
                if cordinate == "tuple":
                    tpl = coordinate_to_tuple(cell_address)
                    result.append(tpl)
                elif cordinate == "address":
                    result.append(cell_address)
            
    return result